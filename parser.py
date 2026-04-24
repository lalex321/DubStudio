"""
Чтение xlsx-файлов серий Netflix. Скопировано из Pivot (consolidator.py) —
только часть, отвечающая за разбор источника. Генерация сводного xlsx
осталась в Pivot.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class Profile:
    name: str
    sheet_name: str
    episode_pattern: str
    total_marker: str
    col_character: int = 0
    col_dialog: int = 1
    col_transcription: int = 2
    col_foreign: int = 3
    col_music: int = 4
    col_burnedin: int = 5
    col_onscreen: int = 6
    col_total: int = 7


PROFILES: dict[str, Profile] = {
    "default": Profile(
        name="Default",
        sheet_name="Word Count Summary",
        episode_pattern=r"(\d+)\s*СЕРИЯ",
        total_marker="TOTAL WORD COUNT BY TEXT CATEGORY",
    ),
}


EpisodeRow = tuple


@dataclass
class EpisodeData:
    number: int
    filename: str
    rows: list[EpisodeRow] = field(default_factory=list)
    total: EpisodeRow | None = None
    show_title: str = ""


# `\b` в Python regex считает `_` частью слова, поэтому используем lookaround
# по буквам — цифры и `_` допустимы как соседние символы.
_NL = r"(?<![A-Za-z])"
_NR = r"(?![A-Za-z])"

_FALLBACK_EPISODE_PATTERNS: list[re.Pattern[str]] = [
    re.compile(p, re.IGNORECASE)
    for p in [
        rf"{_NL}s\d+\s*e(\d+){_NR}",
        # Netflix DUB_SCRIPT: "...MatingSeasonSeason1Episode10TheWolfWedding..."
        # Слово Season может прилегать к названию сериала слева без
        # разделителя, поэтому lookbehind не применяем; right-boundary
        # тоже нет — цифра эпизода прилегает к титулу серии.
        rf"season\d+\s*episode\s*(\d+)",
        rf"{_NL}episode\s*(\d+)",
        rf"{_NL}ep\.?\s*(\d+){_NR}",
        rf"{_NL}e(\d+){_NR}",
        r"(\d+)\s*серия",
        r"серия\s*(\d+)",
        r"^[\s_-]*(\d+)(?=[\s_.\-]|$)",
    ]
]

ACCEPTED_FORMATS_HINT = (
    "Ожидаю форматы: '1 СЕРИЯ ...', 'Episode 1', 'S01E01', "
    "'Ep01', 'E01', '01_...'"
)


def _detect_episode(stem: str, profile_pattern: str) -> tuple[int, str] | None:
    patterns: list[re.Pattern[str]] = []
    try:
        patterns.append(re.compile(profile_pattern, re.IGNORECASE))
    except re.error:
        pass
    patterns.extend(_FALLBACK_EPISODE_PATTERNS)

    for rx in patterns:
        m = rx.search(stem)
        if not m:
            continue
        groups = [g for g in m.groups() if g is not None]
        if not groups:
            continue
        try:
            ep_num = int(groups[0])
        except ValueError:
            continue
        cleaned = rx.sub(" ", stem, count=1)
        return ep_num, cleaned
    return None


def collect_episodes(
    files: list[tuple[str, bytes]],
    profile: Profile,
) -> tuple[dict[int, EpisodeData], list[str]]:
    episodes: dict[int, EpisodeData] = {}
    warnings: list[str] = []

    for fname, blob in files:
        if fname.startswith("~$") or not fname.lower().endswith(".xlsx"):
            warnings.append(f"{fname}: не xlsx, пропущен")
            continue
        stem = Path(fname).stem
        detection = _detect_episode(stem, profile.episode_pattern)
        if detection is None:
            warnings.append(f"{fname}: не нашёл номер эпизода. {ACCEPTED_FORMATS_HINT}")
            continue
        ep_num, _ = detection

        wb = load_workbook(io.BytesIO(blob), data_only=True)
        if profile.sheet_name not in wb.sheetnames:
            warnings.append(f"{fname}: нет листа '{profile.sheet_name}', пропущен")
            continue

        ws = wb[profile.sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            warnings.append(f"{fname}: пустой лист, пропущен")
            continue

        data_rows: list[EpisodeRow] = []
        total_row: EpisodeRow | None = None
        min_width = profile.col_total + 1
        for r in rows[1:]:
            if len(r) < min_width:
                r = r + (None,) * (min_width - len(r))
            if not any(v is not None for v in r):
                continue
            first = str(r[profile.col_character] or "").upper()
            if profile.total_marker in first:
                total_row = r
            elif r[profile.col_character] is not None or any(v for v in r[1:]):
                data_rows.append(r)

        show_title = ""
        if "Project Info" in wb.sheetnames:
            pi = wb["Project Info"]
            for pi_row in pi.iter_rows(values_only=True):
                if not pi_row or not pi_row[0]:
                    continue
                if str(pi_row[0]).strip().upper() == "SHOW TITLE":
                    if len(pi_row) > 1 and pi_row[1] is not None:
                        show_title = str(pi_row[1]).strip()
                    break

        if ep_num in episodes:
            warnings.append(
                f"{fname}: серия {ep_num} уже была (из {episodes[ep_num].filename}), "
                f"беру более позднюю"
            )
        episodes[ep_num] = EpisodeData(
            number=ep_num,
            filename=fname,
            rows=data_rows,
            total=total_row,
            show_title=show_title,
        )

    return episodes, warnings


def derive_show_title(episodes: dict[int, EpisodeData]) -> str:
    titles = sorted(
        {d.show_title.strip() for d in episodes.values() if d.show_title and d.show_title.strip()}
    )
    if not titles:
        return ""
    if len(titles) == 1:
        return titles[0]
    return " / ".join(titles)


_JUNK_TOKENS = {
    "DUB", "SCRIPT", "SUB", "SUBS", "DUBBING",
    "FINAL", "CUT", "APPROVED", "DRAFT", "TEMP", "ROUGH", "FULL",
    "RU", "EN", "ES", "FR", "DE", "IT", "PT", "JA", "KO", "ZH", "PL", "TR",
}
_JUNK_TOKEN_RX = re.compile(r"^(?:v\d+(?:\.\d+)*|r\d+|ver\d*|\d+)$", re.IGNORECASE)


def _is_junk_token(t: str) -> bool:
    if t.upper() in _JUNK_TOKENS:
        return True
    if _JUNK_TOKEN_RX.match(t):
        return True
    return False


def derive_common_name(filenames: list[str], profile: Profile) -> str:
    token_sets_upper: list[set[str]] = []
    reference: list[str] = []

    for i, fname in enumerate(filenames):
        stem = Path(fname).stem
        det = _detect_episode(stem, profile.episode_pattern)
        stripped = det[1] if det is not None else stem
        tokens = [t for t in re.split(r"[\s_\-]+", stripped.strip()) if t]
        if i == 0:
            reference = tokens
        token_sets_upper.append({t.upper() for t in tokens})

    if not reference or not token_sets_upper:
        return ""

    rest = token_sets_upper[1:]
    common = [t for t in reference if all(t.upper() in s for s in rest)]
    common = [t for t in common if not _is_junk_token(t)]
    return " ".join(common).strip()

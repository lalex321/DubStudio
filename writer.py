"""
xlsx-экспорт сводной таблицы проекта. Стили и макет — как в Pivot
(consolidator.py), плюс колонка Actor после Character.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


FONT_NAME = "Arial"
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="305496")
TOTAL_FONT = Font(name=FONT_NAME, bold=True, size=11)
TOTAL_FILL = PatternFill("solid", start_color="FFE699")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
BOLD_SMALL = Font(name=FONT_NAME, bold=True, size=10)
GRAY_FILL = PatternFill("solid", start_color="F2F2F2")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F3864")
METRIC_FONT = Font(name=FONT_NAME, bold=True, size=13, color="1F3864")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header_row(ws, row: int, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER


def _write_sheet(
    ws,
    metric_label: str,
    show_title: str,
    ep_numbers: list[int],
    characters: list[str],
    actor_by_char: dict[str, str],
    values: dict[str, dict[int, int]],
) -> None:
    """
    Пишет лист для одной метрики.
    Колонки: Character | Actor | E01..EN | Total | Episodes
    Row 1 — Show Title, Row 2 — metric_label, Row 3 — заголовки, Row 4+ — данные,
    пустая строка, Row(last+2) — EPISODE TOTAL.
    """
    n_eps = len(ep_numbers)
    last_col = n_eps + 4  # Character + Actor + N episodes + Total + Episodes

    c1 = ws.cell(row=1, column=1, value=show_title or "")
    c1.font = TITLE_FONT
    c1.alignment = CENTER
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)

    c2 = ws.cell(row=2, column=1, value=metric_label)
    c2.font = METRIC_FONT
    c2.alignment = CENTER
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)

    headers = (
        ["Character", "Actor"]
        + [f"E{i:02d}" for i in ep_numbers]
        + ["Total", "Episodes"]
    )
    _style_header_row(ws, 3, headers)

    first_ep_col_letter = get_column_letter(3)
    last_ep_col_letter = get_column_letter(2 + n_eps)
    sum_col = n_eps + 3  # Character(1) Actor(2) E01..EN then Total
    ep_count_col = n_eps + 4

    row = 4
    for name in characters:
        ch_cell = ws.cell(row=row, column=1, value=name)
        ch_cell.font, ch_cell.alignment, ch_cell.border = NORMAL_FONT, LEFT, BORDER

        ac_cell = ws.cell(row=row, column=2, value=actor_by_char.get(name, "") or "")
        ac_cell.font, ac_cell.alignment, ac_cell.border = NORMAL_FONT, LEFT, BORDER

        per_ep = values.get(name, {})
        for i, ep in enumerate(ep_numbers, start=3):
            val = per_ep.get(ep, 0)
            cell = ws.cell(row=row, column=i, value=val if val else None)
            cell.font, cell.alignment, cell.border = NORMAL_FONT, RIGHT, BORDER

        s = ws.cell(
            row=row,
            column=sum_col,
            value=f"=SUM({first_ep_col_letter}{row}:{last_ep_col_letter}{row})",
        )
        s.font, s.alignment, s.border, s.fill = BOLD_SMALL, RIGHT, BORDER, GRAY_FILL

        cnt = ws.cell(
            row=row,
            column=ep_count_col,
            value=f"=COUNT({first_ep_col_letter}{row}:{last_ep_col_letter}{row})",
        )
        cnt.font, cnt.alignment, cnt.border = NORMAL_FONT, RIGHT, BORDER
        row += 1

    last_data_row = row - 1
    # пустая строка-разделитель — защищает EPISODE TOTAL от сортировки автофильтром
    row += 1

    tc = ws.cell(row=row, column=1, value="EPISODE TOTAL")
    tc.font, tc.fill, tc.alignment, tc.border = TOTAL_FONT, TOTAL_FILL, LEFT, BORDER
    # Actor в EPISODE TOTAL — пустая ячейка со стилем
    ac_total = ws.cell(row=row, column=2)
    ac_total.font, ac_total.fill, ac_total.alignment, ac_total.border = (
        TOTAL_FONT, TOTAL_FILL, LEFT, BORDER,
    )
    for i in range(3, len(headers) + 1):
        col = get_column_letter(i)
        cell = ws.cell(row=row, column=i)
        if i != ep_count_col:
            cell.value = f"=SUM({col}4:{col}{last_data_row})"
        cell.font, cell.fill, cell.alignment, cell.border = (
            TOTAL_FONT, TOTAL_FILL, RIGHT, BORDER,
        )

    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{last_data_row}"

    ws.column_dimensions["A"].width = 32  # Character
    ws.column_dimensions["B"].width = 22  # Actor
    for i in range(3, 3 + n_eps):
        ws.column_dimensions[get_column_letter(i)].width = 11
    ws.column_dimensions[get_column_letter(sum_col)].width = 13  # Total
    ws.column_dimensions[get_column_letter(ep_count_col)].width = 11  # Episodes
    ws.row_dimensions[3].height = 32
    ws.freeze_panes = "C4"


def build_actor_report_xlsx(
    show_title: str,
    rows: list[tuple[str, int]],
) -> bytes:
    """Простой отчёт для бухгалтерии: Актёр | Слов (Transcription), отсорт по desc."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Actor Report"

    t = ws.cell(row=1, column=1, value=show_title or "")
    t.font = TITLE_FONT
    t.alignment = CENTER
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    sub = ws.cell(row=2, column=1, value="Actor Report — Transcription Word Count")
    sub.font = METRIC_FONT
    sub.alignment = CENTER
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)

    for i, h in enumerate(["Actor", "Words"], start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, CENTER, BORDER

    start_row = 4
    for idx, (name, words) in enumerate(rows):
        r = start_row + idx
        ac = ws.cell(row=r, column=1, value=name)
        ac.font, ac.alignment, ac.border = NORMAL_FONT, LEFT, BORDER
        wc = ws.cell(row=r, column=2, value=words if words else None)
        wc.font, wc.alignment, wc.border = NORMAL_FONT, RIGHT, BORDER

    last_row = start_row + max(len(rows), 1) - 1
    total_row = last_row + 2
    lc = ws.cell(row=total_row, column=1, value="TOTAL")
    lc.font, lc.fill, lc.alignment, lc.border = TOTAL_FONT, TOTAL_FILL, LEFT, BORDER
    tc = ws.cell(row=total_row, column=2, value=f"=SUM(B{start_row}:B{last_row})")
    tc.font, tc.fill, tc.alignment, tc.border = TOTAL_FONT, TOTAL_FILL, RIGHT, BORDER

    # Excel sort/filter dropdowns on the header row. Range covers only the
    # data block — the blank separator and TOTAL row below stay out of
    # filter scope so TOTAL is never reordered or hidden by a filter.
    ws.auto_filter.ref = f"A3:B{last_row}"

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_project_xlsx(
    show_title: str,
    ep_numbers: list[int],
    characters: list[str],
    actor_by_char: dict[str, str],
    transcription: dict[str, dict[int, int]],
    dialogue: dict[str, dict[int, int]],
) -> bytes:
    wb = Workbook()
    # Transcription — первым (Excel открывает активный лист).
    _write_sheet(
        wb.active,
        metric_label="Transcription Word Count",
        show_title=show_title,
        ep_numbers=ep_numbers,
        characters=characters,
        actor_by_char=actor_by_char,
        values=transcription,
    )
    wb.active.title = "Transcription Summary"

    _write_sheet(
        wb.create_sheet("Dialogue Summary"),
        metric_label="Dialogue Word Count",
        show_title=show_title,
        ep_numbers=ep_numbers,
        characters=characters,
        actor_by_char=actor_by_char,
        values=dialogue,
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
